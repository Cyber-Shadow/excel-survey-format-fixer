#!/bin/python
#23-24/09/2020
#Notes: This program assumes that the questions are numbered 1 to n (can have subquestions abc etc) and the answers are to the right of the question.
#Importing required modules
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter as gcl

#Setting up excel workbooks and worksheets. CHANGE INPUT FILE/SHEET HERE (ENSURE FILES ARE IN THE SAME DIRECTORY)
wbi = load_workbook(filename = 'test.xlsx')
wsi = wbi['INSERT WORKSHEET NAME HERE']
wbo = Workbook()
wso = wbo.active

#Initial variable definitions to stop the program from crying about undefined variables
outputcol = 0
qnum = 0
last = ""
qfirst = 0
wasquestion = False

#Iteration, columns before rows. This goes through every single cell (inefficent) and checks it for the below critera. (counts like this: 1A, 1B, 1C, 2A, 2B, 2C)
for row in range(1,wsi.max_row+1):
    for col in range(1,wsi.max_column+1):
        #For debug:
        print(gcl(col)+str(row) + ' qnum:' + str(qnum) + ' last:' + str(last) + ' qfirst: ' + str(qfirst))

        #Resets counting variables when sorting new entry
        if str(wsi[str(str(gcl(col)) + str(row))].value).find("1.") != -1 and wasquestion is False:
            alist=[]
            print('reset')
            qnum = 1
            qfirst += 1
            outputcol = 1


        #Accounts for sub questions (abc), basically if the question number is the same as the last question don't advance the question number.
        #Note: the question number is used to verify if the cell contains a question, it is possible that if the answer contains the same number as the question it messes up. (fixed by checking if the previous cell was a question)
        if str(qnum -1) in str(last) and wasquestion is False:
            if str(qnum -1 ) in (str(wsi[str(gcl(col)) + str(row)].value)):
                qnum -= 1

        #Checks if the value of the selected cell contains the question number, we can assume this is the question (especially given that the last cell checked was NOT the question)
        if (str(wsi[str(gcl(col)) + str(row)].value).find(str(qnum))) != -1 and wasquestion is False:

            #the question number increases every time a question found so it can
            qnum += 1


            #on the first run through sorting the values, also write the questions in row 1.
            if qfirst == 1:
                #fixes bug that reads empty value as string None. Given that its a real value, write it in the ouput excel sheet
                if (str(wsi[str(gcl(col)) + str(row)].value)) != "None":
                    #print (str(wsi[str(gcl(col)) + str(row)].value))
                    wso[gcl(outputcol) + "1"] = (str(wsi[str(gcl(col)) + str(row)].value))

            #Take the cell to the right of selected (question) cell and write it under the correct column, the worksheet function looks like ws['A1']
            print (str(wsi[str(gcl(col)) + str(row)].value))
            wso[gcl(outputcol) + str(qfirst + 1)] = (str(wsi[str(gcl(col+1)) + str(row)].value))

            #signal that on the next write, write it to the right of the one that was just written.
            outputcol += 1

            #writes variable "last" which enables the program to check if the last question has the same question number as the current question (used above)
            last = (str(wsi[str(gcl(col)) + str(row)].value))

            #if this cell is a question, then the next cell (the answer) must NOT be treated like a question. Hence, it only treats the cell as a question if the last cell was NOT a question.
            wasquestion = True
        else:
            wasquestion = False

#saves the workbook/excel file generated by this program
wbo.save('output.xlsx')
